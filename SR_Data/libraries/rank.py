from numpy import datetime64
from .pres_template import *

import pandas as pd
# import investpy
from investiny import historical_data, search_assets
import time
# from os.path import exists, getmtime
from openpyxl.styles import Font, Alignment

from .graphics import *
from .getdata import *
from .coins import *
from .countries import *


def read_csv_last_date(file_name):
    if exists(file_name):
        df = pd.read_csv(file_name)
        last_date_ok = df['Date'].iloc[-10]
        df = df[:-10]
    else:
        df = pd.DataFrame()
        # last_date_ok = '1979-12-01'
        last_date_ok = '2009-12-01'

    from_date = f'{last_date_ok[8:10]}/{last_date_ok[5:7]}/{last_date_ok[0:4]}'

    return df, last_date_ok, from_date


def save_new_data(file_name, df, df_new, last_date_ok):
    df_new['Date'] = df_new['Date'].astype(str)
    df = df.append(df_new[df_new.Date > last_date_ok]).reset_index(drop=True)
    df.to_csv(file_name, index=False, line_terminator='\n') #, compression='zip')

    return df


update_rank = False
def set_update_rank(st):
    global update_rank
    update_rank = st


def rank_load_data(country_name, index_name):

    file_name = f"data/rank/{index_name.replace('/', '_')}.csv"

    if check_date_file(file_name, same='month') and not update_rank:
        df = pd.read_csv(file_name) #, compression='zip')
    else:
        time.sleep(2)  # to avoid been blocked

        df, last_date_ok, from_date = read_csv_last_date(file_name)

        # df_new = get_index_historical_data(
        #     index=index_name, country=country_name,
        #     from_date=from_date,
        #     to_date='31/12/2029').reset_index()

        search_results = search_assets(query=index_name, limit=1, type="Index") #, exchange="NASDAQ")
        investing_id = int(search_results[0]["ticker"]) # Assuming the first entry is the desired one (top result in Investing.com)
        df_new = historical_data(
            investing_id=investing_id,
            from_date = f'{from_date[3:5]}/{from_date[0:2]}/{from_date[6:10]}',
            to_date='12/31/2029')
        df_new = pd.DataFrame(df_new).reset_index().rename(columns={'date':'Date', 'open':'Open', 'high':'High', 'low':'Low', 'close':'Close', 'volume':'Volume'})
        df_new['Date'] = pd.to_datetime(df_new['Date'])

        df = save_new_data(file_name, df, df_new, last_date_ok)

    df['Date'] = df['Date'].astype(datetime64)

    return df


def rank_load_currency(country_name):
    curr_code = country_to_currency(country_name)

    if curr_code == 'USD':
        return 0

    file_name = f"data/rank/USD_{curr_code}.csv"
    currency_cross=f'USD/{curr_code}'

    if check_date_file(file_name, same='month') and not update_rank:
        df = pd.read_csv(file_name) #, compression='zip')
    else:
        time.sleep(2)  # to avoid been blocked

        df, last_date_ok, from_date = read_csv_last_date(file_name)

        # df_new = get_currency_cross_historical_data(
        #     currency_cross=currency_cross,
        #     from_date=from_date,
        #     to_date='31/12/2029').reset_index()

        search_results = search_assets(query=currency_cross, limit=10, type='FX')
        for item in search_results:
            if item['symbol'] == currency_cross:
                investing_id = item['ticker']
                break
        else:
            investing_id = int(search_results[0]['ticker']) # Assuming the first entry is the desired one (top result in Investing.com)
        df_new = historical_data(
            investing_id=investing_id,
            from_date = f'{from_date[3:5]}/{from_date[0:2]}/{from_date[6:10]}',
            to_date='12/31/2029')
        df_new = pd.DataFrame(df_new).reset_index().rename(columns={'date':'Date', 'open':'Open', 'high':'High', 'low':'Low', 'close':'Close', 'currency':'Currency'})
        df_new['Date'] = pd.to_datetime(df_new['Date'])

        df = save_new_data(file_name, df, df_new, last_date_ok)

    df['Date'] = df['Date'].astype(datetime64)

    return df


def draw_rank(fig, annot, countries, values, br_pos, pt_mode=None, show_perc=False, debug=None):
    val_min = values.min()
    val_max = values.max()
    val_range = val_max - val_min

    w, bh  = RANK_BAR_W, RANK_BAR_H
    dx = RANK_BAR_DX

    y0 = RANK_Y + 10

    # Referencial variável
    if val_min > 0: val_min = 0
    if val_max < 0: val_max = 0
    val_range = val_max - val_min

    if val_min < 0:
        if val_max <= 0:  # só valores para baixo
            y0 += 80  # desce y0 para caber bandeira em cima
        else:
            if bh * val_max/val_range < 90:  # barra acima NÃO SUPERA bandeira
                if bh * val_min/val_range < -90:  # barra abaixo SUPERA bandeira
                    y0 += 80
                else:
                    bh -= 80  # só mexe no tamanho se barras não superarem bandeiras

    y0 += bh * val_min / val_range

    size = len(countries)

    for i in range(size):
        val = values[i]

        x0 = RANK_X + i * dx
        x1 = x0 + w

        y1 = y0 - bh * val / val_range

        if val >= 0:
            color = BLOCK_BG_COLOR_MID_UP
            flag_circle_color = FLAG_CIRCLE_COLOR_UP
        else:
            color = BLOCK_BG_COLOR_MID_DN
            flag_circle_color = FLAG_CIRCLE_COLOR_DN

        if TEMPLATE == 'JP_MERC_FIN':
            txt_color = FOCUS_FOOT_COLOR
        else:
            txt_color = flag_circle_color

        if TEMPLATE == 'JP_MERC_FIN_4':
            txt_color = '#FFFFFF'

        draw_rectangle(fig, x0, y0, x1, y1, color)

        # Positions
        val_y  = y1 - 4
        val_yanchor  = 'bottom'
        flag_yanchor = 'top'
        img_file = BLOCK_VERT_DN_IMG
        dir = 1
        if val < 0:
            val_y  = y1 + 4
            val_yanchor = 'top'
            flag_yanchor = 'bottom'
            img_file = BLOCK_VERT_UP_IMG
            dir = -1

        val_txt = get_value_txt(val, bold=True)
        if show_perc:
            val_txt += '%'
        if  gvar.get('val_suffix'):
            val_txt += gvar['val_suffix']
        add_annot(annot, x0 + w/2, val_y, val_txt,
        txt_color, 24, # FOCUS_FOOT_FONT_SIZE,
        xanchor='center', yanchor=val_yanchor)

        hei = w / BLOCK_VERT_W * BLOCK_VERT_H
        # if img_file:
        #     add_image(fig, img_file, x0, y0 - dir, w, h,
        #             xanchor='left', yanchor=flag_yanchor)

        if not gvar.get('JP_Ibope'):
            if TEMPLATE == 'INVEST_NEWS_BLACK':
                draw_circle_line(fig, x0+ 8, y0+dir*11, x1- 8, y0+dir*(11+w-16), flag_circle_color, width=3)
            else:
                draw_circle(fig, x0+ 8, y0+dir*11, x1- 8, y0+dir*(11+w-16), flag_circle_color)
            draw_circle(fig, x0+12, y0+dir*15, x1-12, y0+dir*(15+w-24), 'white')

        country = countries[i]
        isoa3 = country_to_isoa3(country)

        if gvar.get('JP_Ibope', False):
            img_file = f"images/jp_ibope_icons/{country.upper().replace(' ', '_')}.png"
        else:
            img_file = f'images/flags/{isoa3}.png'

        imgx, imgy = x0+106/568*w, y0 - dir + dir*(134/568*w)
        imgw, imgh = 356/568*w, 356/568*w,
        if  gvar.get('TEMPLATE') == 'JP_IBOPE':
            imgx -= 6
            imgy -= 8
            imgw, imgh = 64, 64
        add_image(fig, img_file, imgx, imgy, imgw, imgh,
                xanchor='left', yanchor=flag_yanchor)

        if not pt_mode:
            if country == 'european union':
                country_pt = 'União<br>Européia'
            else:
                pos = idx_country_names.index(country)
                country_pt = idx_country_names_pt[pos]
        else:
            country_pt = country_to_portuguese(country)

        font_size = FOCUS_FOOT_FONT_SIZE
        country_len = max([len(cline) for cline in country_pt.split('<br>')])
        if country_len > 8:  # -6% for each letter
            font_size *= 1 - (country_len - 8) * 0.06
            if (font_size < 8):
                font_size = 8

        # Posição do país
        country_pos = size - i
        if country_pt == 'Brasil' and i == 0:
            country_pos = br_pos

        y_ctnam = y0 + dir * (hei - 2)
        y_ctpos = y_ctnam + dir * (1 + ('<br>' in country_pt)) * (font_size + 4)

        add_annot(annot, x0 + w/2, y_ctnam, f'<b>{country_pt}</b>',
            txt_color, font_size, xanchor='center', yanchor=flag_yanchor)

        add_annot(annot, x0 + w/2, y_ctpos, f'<b>{country_pos}º</b>',
            txt_color, FOCUS_FOOT_FONT_SIZE, xanchor='center', yanchor=flag_yanchor)

    x0 = RANK_X - 30
    x1 = RANK_X + size * dx
    draw_line(fig, x0, y0, x1, y0, RANK_LINE_COLOR, 2)
    set_graph_center((x0 + x1) / 2)


def adjust_index(index_name, val1, date1, val2, date2):
    # bursatil: https://en.wikipedia.org/wiki/%C3%8Dndice_Burs%C3%A1til_de_Capitalizaci%C3%B3n
    # 2021-03-12, 2645283.75
    # 2021-03-15     2668.31
    if index_name.lower() == 'bursatil':
        for cut_0_date in [dt(2021, 3, 15), dt(2018, 11, 6), dt(2017, 10, 9)]:
            if date1 < cut_0_date: val1 /= 1_000
            if date2 < cut_0_date: val2 /= 1_000
    return val1, val2

def adjust_currency(country, val1, date1, val2, date2):
    if country.lower() == 'venezuela':
        cut_0_date = dt(2021, 9, 27)
        if date1 < cut_0_date: val1 /= 1_000_000
        if date2 < cut_0_date: val2 /= 1_000_000
    return val1, val2


# Vários buracos antes de Dez/2017
def show_rank(fig, annot, symbol, date_ini, date_end, dend_type, table_name, qtd=10, debug=None):

    # if debug: debug.value = 'WAIT ...'

    if 'pior' in symbol.lower():
        qtd = -qtd

    countries = []
    values = []
    index_names, countries_pt, curr_codes = [],[],[]
    vals_ini, vals_end = [],[]
    val_ini_dates, val_end_dates = [],[]
    USD_ini_dates, USD_end_dates = [],[]

    # if 'AmLat' in symbol:
    #     idx_group = idx_latin_america.copy()
    # else:
    #     idx_group = idx_country_names.copy()

    # if '-ven' in symbol.lower():
    #     idx_group.remove('Venezuela')

    idx_group = idx_country_names.copy()
    data = {'xaxis':pd.Series(idx_group), 'yaxis':pd.Series(idx_group)}
    filter_countries(data)
    idx_group = data['xaxis'].to_list()
    if 'united states' in idx_group:
        idx_group.append('united states 2')

    if 'MOEDAS' in symbol:
        while 'United States' in idx_group:
            idx_group.remove('United States')

    for country in idx_group:
        pos = idx_country_names.index(country)
        index_name = idx_names[pos]
        country_ok = country
        country = country_adjust_name(country)
        curr_code = country_to_currency(country)

        if curr_code == 'USD' and 'MOEDAS' in symbol:
            continue

        df = rank_load_data(country, index_name)
        dfc = rank_load_currency(country)

        # Valores default para evitar erro
        to_USD_ini, USD_ini_date = 1, date_ini
        to_USD_end, USD_end_date = 1, date_end

        # Loop para ajustar datas entre Bolsa e Moeda
        date_ini_ok = date_ini
        loop_cnt, no_value = 0, False
        date_ini_bigger = False

        while True:
            if not 'MOEDAS' in symbol:
                if date_ini_ok >= df.Date.iloc[0] and date_ini_bigger == False:
                    val_ini, val_ini_date = df[df.Date <= date_ini_ok].iloc[-1][['Close', 'Date']]
                else:  # pega primeira data disponivel
                    val_ini, val_ini_date = df[df.Date >= date_ini_ok].iloc[ 0][['Close', 'Date']]

                # val_end, val_end_date = df[df.Date <= date_end].iloc[-1][['Close', 'Date']]
                val_end, val_end_date = df[df.Date <= date_end].iloc[-1][[dend_type, 'Date']]

                val_ini, val_end = adjust_index(index_name,
                    val_ini, date_ini, val_end, date_end)
            else:
                val_ini_date = date_ini_ok
                val_end_date = date_end

            # To US$
            if type(dfc) == pd.DataFrame and ('_USD' in symbol or 'MOEDAS' in symbol):
                if val_ini_date >= dfc.Date.iloc[0]:
                    to_USD_ini, USD_ini_date = dfc[dfc.Date <= val_ini_date].iloc[-1][['Close', 'Date']]
                else:
                    to_USD_ini, USD_ini_date = dfc[dfc.Date >= val_ini_date].iloc[ 0][['Close', 'Date']]

                # to_USD_end, USD_end_date = dfc[dfc.Date <= val_end_date].iloc[-1][['Close', 'Date']]
                dend_type_ok = dend_type
                if   dend_type_ok == 'High': dend_type_ok = 'Low'
                elif dend_type_ok == 'Low' : dend_type_ok = 'High'
                to_USD_end, USD_end_date = dfc[dfc.Date <= val_end_date].iloc[-1][[dend_type_ok, 'Date']]

                to_USD_ini, to_USD_end = adjust_currency(country,
                    to_USD_ini, date_ini, to_USD_end, date_end)

                # Aparentemente só acontece com a Venezuela
                if to_USD_ini == 0 or to_USD_end == 0:
                    no_value = True
                    break

                if '_USD' in symbol:
                    val_ini /= to_USD_ini
                    val_end /= to_USD_end
                else:
                    val_ini = 1 / to_USD_ini
                    val_end = 1 / to_USD_end

            if val_ini_date < USD_ini_date:
                date_ini_ok = USD_ini_date
                loop_cnt += 1
                if loop_cnt < 5:
                    continue

            # Aceita até 15 dias de intervalo entre bolsa e dólar
            if '_USD' in symbol and curr_code != 'USD' and not 'MOEDAS' in symbol:
                if val_ini_date - USD_ini_date >= np.timedelta64(15, 'D'):
                    date_ini_ok = dfc[dfc.Date >= val_ini_date].iloc[ 0]['Date']
                    date_ini_bigger = True
                    loop_cnt += 1
                    if loop_cnt < 5:
                        continue

            break

        # Aparentemente só acontece com a Venezuela
        if no_value:
            continue

        val_diff = val_end - val_ini
        val_perc = 0
        if val_ini > 0:
            val_perc = val_diff / val_ini * 100

        pos = idx_country_names.index(country_ok)
        country_pt = idx_country_names_pt[pos].replace('<br>',' ')
        if 'MOEDAS' in symbol and curr_code == 'EUR':
            country_ok = 'european union'
            country_pt = 'União Européia'

        index_names.append(index_name)
        curr_codes.append(curr_code)
        countries.append(country_ok)
        countries_pt.append(country_pt)
        vals_ini.append(val_ini)
        vals_end.append(val_end)
        values.append(val_perc)

        val_ini_dates.append(val_ini_date.strftime("%d/%m/%Y"))
        val_end_dates.append(val_end_date.strftime("%d/%m/%Y"))
        USD_ini_dates.append(USD_ini_date.strftime("%d/%m/%Y"))
        USD_end_dates.append(USD_end_date.strftime("%d/%m/%Y"))

    # Ordena as bolsas ou moedas
    dfr = pd.DataFrame({
        'rank':[0] * len(values),
        'idx_name':index_names,
        'curr_code':curr_codes,
        'country':countries,
        'country_pt':countries_pt,
        'val_ini':vals_ini,
        'val_end':vals_end,
        'perc':values,

        # 'val_ini_dates':val_ini_dates,
        # 'USD_ini_dates':USD_ini_dates,
        # 'val_end_dates':val_end_dates,
        # 'USD_end_dates':USD_end_dates,

        }) # .sort_values('perc', ascending=(qtd > 0)).reset_index(drop=True)

    if not 'MOEDAS' in symbol and '_USD' in symbol:
        dfr['val_ini_dates'] = val_ini_dates
        dfr['USD_ini_dates'] = USD_ini_dates
        dfr['val_end_dates'] = val_end_dates
        dfr['USD_end_dates'] = USD_end_dates
        cols_to_rename = ['date_ini', 'USD_ini', 'date_end', 'USD_end']
    elif not 'MOEDAS' in symbol:
        dfr['val_ini_dates'] = val_ini_dates
        dfr['val_end_dates'] = val_end_dates
        cols_to_rename = ['date_ini', 'date_end']
    else:
        dfr['USD_ini_dates'] = USD_ini_dates
        dfr['USD_end_dates'] = USD_end_dates
        cols_to_rename = ['USD_ini', 'USD_end']

    dfr = dfr.sort_values('perc', ascending=(qtd > 0)).reset_index(drop=True)

    if 'MOEDAS' in symbol:
        dfr.drop_duplicates(subset='curr_code', ignore_index=True, inplace=True)

    dfr['rank'] = list(range(dfr.shape[0], 0, -1))
    dfr['rank'] = dfr['rank'].astype(str) + 'º'

    cols_to_del = ['country', 'curr_code']
    header=['Ranking', 'Bolsa', 'País',
            date_ini.strftime("%d/%m/%Y"), date_end.strftime("%d/%m/%Y"), 'Var.%'] + cols_to_rename
    if 'MOEDAS' in symbol:
        header[1] = 'Cód. Moeda'
        cols_to_del[1] = 'idx_name'

    # Tenta salvar arquivo (se estiver aberto no excel, gera erro)
    try:
        df_xl = dfr.drop(cols_to_del, 1).sort_values('perc', ascending=(qtd < 0))
        writer = pd.ExcelWriter(f'tables/{table_name}.xlsx', engine='openpyxl')
        df_xl.to_excel(writer, sheet_name='Bolsas', index=False, header=header,)

        # https://openpyxl.readthedocs.io/en/stable/styles.html
        # workbook = openpyxl.Workbook()
        ws = writer.sheets['Bolsas']
        colw = [12, 26, 15, 14, 14, 12]
        for pos in range(6):
            col_let = chr(65+pos)
            ws[f'{col_let}1'].font = Font(name='Calibri', bold=True)
            ws[f'{col_let}1'].alignment = Alignment(horizontal="center", vertical="bottom")
            # column_length = max(dfr[column].astype(str).map(len).max(), len(column))
            colf = ws.column_dimensions[ col_let ]
            colf.width = colw[pos]
            if pos >= 3:
                if pos < 5 and 'MOEDAS' in symbol:
                    colf.number_format = u'#,##0.000000'
                else:
                    colf.number_format = u'#,##0.00'
        for cell in ws['A']:
            cell.alignment = Alignment(horizontal="center")

        for pos in range(6, 10):
            col_let = chr(65+pos)
            colf = ws.column_dimensions[ col_let ]
            colf.width = 15
            for cell in ws[col_let]:
                cell.alignment = Alignment(horizontal="center")

        writer.save()
    except:
        pass  # Arquivo já existe e está aberto

    rank_tot = dfr.country.size

    # Testa se Brazil está na lista
    d_brazil = dfr.country[dfr.country == 'brazil']
    br_pos = None
    if d_brazil.size > 0:
        br_pos = d_brazil.index[0]

    # Posição do Brasil
    dfr_br = 0
    if br_pos != None and rank_tot - br_pos > abs(qtd):
        dfr_br = dfr.iloc[br_pos:br_pos+1].copy()

    # Limita quantidade de ítens no gráfico para exibição
    first = dfr.shape[0] - abs(qtd)
    if first > 0:
        dfr = dfr[first:].reset_index(drop=True)

    # Coloca Brazil na lista
    if type(dfr_br) != int:
        # Gambiarra para ficar em 10 no máximo (Invest_News_Black)
        if TEMPLATE == 'INVEST_NEWS_BLACK' and dfr.shape[0] > 9:
            dfr = pd.concat([dfr_br, dfr[1:]], ignore_index=True)
        else:
            dfr = pd.concat([dfr_br, dfr], ignore_index=True)

    if br_pos != None:
        br_pos = rank_tot-br_pos
    draw_rank(fig, annot, dfr.country, dfr.perc, br_pos, show_perc=SHOW_PERC, debug=debug)

    # if debug: debug.value = ''

    return rank_tot


country_filter = ['ALL']

def set_country_filter(filter):
    country_filter[0] = filter

def filter_countries(data):
    # Lista inteira original
    xaxis = data['xaxis'].tolist()
    if not gvar.get('JP_Ibope', False):  # Use get() with default False
        xaxis = [x.lower() for x in xaxis]
    yaxis = data['yaxis'].tolist()

    # Pega filtro dos países
    df_cf = pd.read_csv(f'data/filter/{country_filter[0]}.csv')
    countries = df_cf.countries.tolist()

    # Países para adicionar e subtrair
    countries_to_add, countries_to_sub = [],[]
    for country in countries:
        country = country.strip().lower()
        if country[0] == '-':
            countries_to_sub.append(country[1:].strip())
        else:
            countries_to_add.append(country)

    # Inclui todos os países se não houver nenhum na lista ('ALL')
    if len(countries_to_add) == 0:
        countries_to_add = xaxis[:]

    # Remove os países com '-' no inicio
    for country in countries_to_sub:
        if country in countries_to_add:
            countries_to_add.remove(country)

    # Salva no arquivo de erros os países não encontrados
    if len(countries_to_add) != 0:
        error_str = ('Países não encontrados:\n')
        for country in countries_to_add:
            if country not in xaxis:
                error_str += f'[{country}]\n'
        save_error(error_str)

    # Nova lista ajustada
    xaxis_ok, yaxis_ok = [],[]
    for i in range(len(xaxis)):
        if xaxis[i] in countries_to_add:
            xaxis_ok.append(xaxis[i])
            yaxis_ok.append(yaxis[i])

    # Altera em data
    data['xaxis'] = pd.Series(xaxis_ok)
    data['yaxis'] = pd.Series(yaxis_ok)


def show_rank_dig(fig, annot, symbol, qtd=10, debug=None):
    data, save_file_name = get_symbol_data(symbol)
    if gvar['ERROR_STATE']:
        return None, None

    save_file_name += f'_{country_filter[0]}'

    # Pode vir no arquivo a ordem dos dados (normal ou inversa)
    filter_countries(data)

    # Ordena valores
    dfo = pd.DataFrame({'x':data['xaxis'], 'y':data['yaxis']})
    ascending = (data['type'] == 'ranki')
    dfo = dfo.sort_values('y', ascending=ascending).reset_index(drop=True)
    data['xaxis'] = dfo.x
    data['yaxis'] = dfo.y

    # Posição do Brasil
    d_country = data['xaxis']
    rank_tot = d_country.size

    # Testa se Brazil está na lista
    d_brazil = d_country[d_country.str.lower() == 'brazil']
    br_pos = None
    if d_brazil.size > 0:
        br_pos = d_brazil.index[0] + 1

    # Quantidade para ser exibida vem de configuração
    if data['quant'] != 0:
        qtd = data['quant']

    d_country = (data['xaxis'][:qtd])[::-1].reset_index(drop=True)
    d_perc = (data['yaxis'][:qtd])[::-1].reset_index(drop=True)

    # if rank_tot - br_pos > abs(qtd):
    if br_pos != None and br_pos > abs(qtd):
        # Gambiarra para ficar em 10 no máximo (Invest_News_Black)
        if TEMPLATE == 'INVEST_NEWS_BLACK' and d_country.size > 9:
            d_country = pd.concat([pd.Series(['brazil']), d_country[1:]], ignore_index=True)
            d_perc = pd.concat([pd.Series([data['yaxis'][br_pos-1]]), d_perc[1:]], ignore_index=True)
        else:
            d_country = pd.concat([pd.Series(['brazil']), d_country], ignore_index=True)
            d_perc = pd.concat([pd.Series([data['yaxis'][br_pos-1]]), d_perc], ignore_index=True)

    show_perc = data.get('show_perc', SHOW_PERC)

    draw_rank(fig, annot, d_country, d_perc, br_pos, pt_mode=True, show_perc=show_perc, debug=debug)  # rank_tot-br_pos
    show_extra_labels(fig, annot, symbol, data, 20, None)

    subtit = data.get('subtit', '')
    filter_text = country_subtit[country_filters.index(country_filter[0])]
    if subtit != '':
        subtit = subtit.replace('{quantity}', f'{rank_tot}')
        subtit = subtit.replace('{filter_text}', f'{filter_text}')
    else:
        subtit = f'{rank_tot} {filter_text}'

    return save_file_name, subtit
